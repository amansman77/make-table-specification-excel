from ast import If
import pymysql
import openpyxl
import shutil
import datetime
import config

def showDatabases(cur):
    sql ='''
    SHOW DATABASES
    '''

    cur.execute(sql)
    rows = cur.fetchall()

    print('\tTable count: {:,}'.format(len(rows)))
    return rows

def findTable(cur, database_name):
    sql ='''
    SELECT TABLE_NAME, TABLE_COMMENT
    FROM INFORMATION_SCHEMA.TABLES
    WHERE TABLE_SCHEMA = %s
    '''

    cur.execute(sql, [database_name])
    rows = cur.fetchall()

    print('\tTable count: {:,}'.format(len(rows)))
    return rows

def findColumn(cur, database_name, table_name):
    sql ='''
    SELECT ORDINAL_POSITION, COLUMN_NAME, DATA_TYPE, IFNULL(NUMERIC_PRECISION, CHARACTER_MAXIMUM_LENGTH) AS DATA_LENGTH, IS_NULLABLE, COLUMN_TYPE, COLUMN_KEY, EXTRA, COLUMN_COMMENT
    FROM INFORMATION_SCHEMA.COLUMNS
    WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
    ORDER BY ORDINAL_POSITION ASC
    '''

    cur.execute(sql, [database_name, table_name])
    rows = cur.fetchall()

    print('\tColumn count: {:,}'.format(len(rows)))
    return rows

def findIndex(cur, database_name, table_name):
    sql ='''
    SELECT INDEX_NAME, COLUMN_NAME
    FROM information_schema.STATISTICS
    WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s AND INDEX_NAME = 'PRIMARY'
    UNION ALL
    (SELECT INDEX_NAME, GROUP_CONCAT(COLUMN_NAME SEPARATOR ', ') as COLUMN_NAME
    FROM information_schema.STATISTICS
    WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s AND INDEX_NAME != 'PRIMARY'
    GROUP BY INDEX_NAME
    ORDER BY NON_UNIQUE ASC, INDEX_NAME ASC, SEQ_IN_INDEX ASC)
    '''

    cur.execute(sql, [database_name, table_name, database_name, table_name])
    rows = cur.fetchall()

    print('\tIndex count: {:,}'.format(len(rows)))
    return rows
    
def generateCreateTableSql(table_row, column_rows, index_rows):
    create_table = 'CREATE TABLE `{table_name}`'.format(table_name=table_row['TABLE_NAME'])

    column_sql = ''
    for i, column_row in enumerate(column_rows):
        column_name_sql = "`{column_name}`".format(
            column_name=column_row['COLUMN_NAME']
        )
        data_type_sql = column_row['DATA_TYPE'].upper()
        if column_row['DATA_LENGTH'] is not None:
            data_type_sql = data_type_sql + '({data_length})'.format(data_length=column_row['DATA_LENGTH'])
        if 'unsigned' in column_row['COLUMN_TYPE']:
            data_type_sql = data_type_sql + ' UNSIGNED'
        null_sql = ''
        if column_row['IS_NULLABLE'] == 'NO':
            null_sql = 'NOT NULL'
        else:
            null_sql = 'NULL DEFAULT NULL'
        if 'auto_increment' in column_row['EXTRA']:
            null_sql = null_sql + ' AUTO_INCREMENT'
        pk_sql = ''
        index_sql = ''
        for i, index_row in enumerate(index_rows):
            if index_row['INDEX_NAME'] == 'PRIMARY':
                pk_sql = '    PRIMARY KEY (`{column_name}`) USING BTREE'.format(column_name=index_row['COLUMN_NAME'])
            else:
                if len(index_sql) > 0:
                    index_sql = index_sql + ',\n'
                index_sql = index_sql + '    INDEX `{index_name}` ({column_name}) USING BTREE'.format(
                    index_name=index_row['INDEX_NAME'],
                    column_name=index_row['COLUMN_NAME']
                )

        if len(column_sql) > 0:
            column_sql = column_sql + ',\n'
        column_sql = column_sql + "    {column_name_sql} {data_type_sql} {null_sql}".format(
            column_name_sql=column_name_sql,
            data_type_sql=data_type_sql,
            null_sql=null_sql
        )
        if column_row['COLUMN_COMMENT'] is not None and len(column_row['COLUMN_COMMENT']) > 0:
            column_sql = column_sql + " COMMENT '{column_comment}'".format(column_comment=column_row['COLUMN_COMMENT'])

    # 문서상 띄어쓰기가 기입되어 들여쓰기를 제거함
    pre_index_sql = '''{create_table} (
{column_sql},
{pk_sql}'''.format(
    create_table=create_table,
    column_sql=column_sql,
    pk_sql=pk_sql
)
    post_index_sql = '''
)
COMMENT='{table_comment}'
ENGINE=InnoDB;'''.format(
    table_comment=table_row['TABLE_COMMENT']
)

    sql = pre_index_sql
    if len(index_sql) > 0:
        sql = sql + ',\n' + index_sql
    sql = sql + post_index_sql

    return sql

def addSheet(wb, table_row, column_rows, index_rows, craete_table_sql):
    ws = wb.copy_worksheet(wb[cell_config['template_sheet_name']])
    ws.title = table_row['TABLE_NAME']
    ws[cell_config['TABLE_ENGLISH_NAME']] = table_row['TABLE_NAME']
    ws[cell_config['TABLE_KOREAN_COMMENT']] = table_row['TABLE_COMMENT']
    ws[cell_config['TABLE_COMMENT']] = table_row['TABLE_COMMENT']

    start_column_idx = cell_config['START_COLUMN_INDEX']
    start_index_idx = cell_config['START_INDEX_INDEX']
    # insert row의 개념이 이하 row를 move하는 개념인데 move시에 병합된 cell이 깨지는 현상 존재
    # ws.insert_rows(start_column_idx + 1, len(column_rows))
    for i, column_row in enumerate(column_rows):
        row_index = start_column_idx + i
        ws[cell_config['COLUMN_NUMBER'] + str(row_index)] = column_row['ORDINAL_POSITION']
        ws[cell_config['COLUMN_NAME'] + str(row_index)] = column_row['COLUMN_NAME']
        ws[cell_config['DATA_TYPE'] + str(row_index)] = column_row['DATA_TYPE']
        ws[cell_config['DATA_LENGTH'] + str(row_index)] = column_row['DATA_LENGTH']
        ws[cell_config['IS_NULLABLE'] + str(row_index)] = column_row['IS_NULLABLE']
        ws[cell_config['COLUMN_KEY'] + str(row_index)] = column_row['COLUMN_KEY']
        ws[cell_config['EXTRA'] + str(row_index)] = column_row['EXTRA']

    for i, index_row in enumerate(index_rows):
        row_index = start_index_idx + i
        ws[cell_config['INDEX_NUMBER'] + str(row_index)] = (i + 1)
        ws[cell_config['INDEX_NAME'] + str(row_index)] = index_row['INDEX_NAME']
        ws[cell_config['INDEX_COLUMN_NAME'] + str(row_index)] = index_row['COLUMN_NAME']
        
        if index_row['INDEX_NAME'] == 'PRIMARY':
            ws[cell_config['INDEX_TYPE'] + str(row_index)] = '로컬 (PK)'
        
        ws[cell_config['INDEX_COLUMN_NAME_2'] + str(row_index)] = index_row['COLUMN_NAME']

    ws[cell_config['CREATE_TABLE_SQL']] = craete_table_sql

if __name__ == '__main__':
    file_config = config.FILE
    database_config = config.DATABASE
    cell_config = config.CELL_INFO

    file_path = file_config['template_full_path']
    file_path_output = file_path[:file_path.rindex('/')] + file_config['file_name']

    today = datetime.datetime.now()
    create_date = today.strftime('%Y%m%d')
    file_path_output = file_path_output.replace('{{date}}', create_date)

    database_name = database_config['database_name']
    file_path_output = file_path_output.replace('{{database_name}}', database_name)

    shutil.copyfile(file_path, file_path_output)

    connection = pymysql.connect(
        host=database_config['host'],
        port=database_config['port'],
        user=database_config['user'],
        passwd=database_config['passwd'],
        database=database_name,
        charset=database_config['charset']
    )
    cursor = connection.cursor(pymysql.cursors.DictCursor)

    table_rows = findTable(cursor, database_name)

    wb = openpyxl.load_workbook(filename = file_path_output)
    for table_row in table_rows:
        column_rows = findColumn(cursor, database_name, table_row['TABLE_NAME'])
        index_rows = findIndex(cursor, database_name, table_row['TABLE_NAME'])
        craete_table_sql = generateCreateTableSql(table_row, column_rows, index_rows)
        # print(table_row)
        # print(column_rows)

        addSheet(wb, table_row, column_rows, index_rows, craete_table_sql)

    wb.remove(wb[cell_config['template_sheet_name']])
    wb.save(file_path_output)
    